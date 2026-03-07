import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from sqlalchemy.orm import Session
from app.models.product import Product
from app.models.stock_movement import StockMovement
from datetime import datetime


def export_products_excel(db: Session) -> bytes:
    products = db.query(Product).filter(Product.is_active == True).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory Report"

    # Header styling
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")

    headers = [
        "SKU", "Product Name", "Category", "Supplier",
        "Cost Price", "Selling Price", "Current Stock",
        "Reorder Level", "Stock Value", "Status", "Expiry Date"
    ]

    # Set column widths
    col_widths = [15, 25, 15, 15, 12, 13, 14, 14, 12, 12, 14]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    # Write data rows
    for row, product in enumerate(products, 2):
        stock_value = round(product.current_stock * product.unit_price, 2)

        if product.current_stock == 0:
            status = "Out of Stock"
        elif product.current_stock <= product.reorder_level:
            status = "Low Stock"
        else:
            status = "In Stock"

        row_data = [
            product.sku,
            product.name,
            product.category.name if product.category else "N/A",
            product.supplier.name if product.supplier else "N/A",
            product.unit_price,
            product.selling_price,
            product.current_stock,
            product.reorder_level,
            stock_value,
            status,
            str(product.expiry_date) if product.expiry_date else "N/A"
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")

            # Color code status
            if col == 10:
                if status == "Out of Stock":
                    cell.font = Font(color="DC2626", bold=True)
                elif status == "Low Stock":
                    cell.font = Font(color="D97706", bold=True)
                else:
                    cell.font = Font(color="059669", bold=True)

        # Alternate row colors
        if row % 2 == 0:
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = PatternFill(
                    start_color="F5F3FF", end_color="F5F3FF", fill_type="solid"
                )

    # Summary row
    summary_row = len(products) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=11)
    ws.cell(row=summary_row + 1, column=1, value="Total Products:")
    ws.cell(row=summary_row + 1, column=2, value=len(products))
    ws.cell(row=summary_row + 2, column=1, value="Total Stock Value:")
    ws.cell(row=summary_row + 2, column=2, value=round(
        sum(p.current_stock * p.unit_price for p in products), 2
    ))
    ws.cell(row=summary_row + 3, column=1, value="Generated On:")
    ws.cell(row=summary_row + 3, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M"))

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def export_products_pdf(db: Session) -> bytes:
    products = db.query(Product).filter(Product.is_active == True).all()

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=0.5 * inch,
        leftMargin=0.5 * inch,
        topMargin=0.5 * inch,
        bottomMargin=0.5 * inch
    )

    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph(
        "<b>Smart Inventory Management — Inventory Report</b>",
        styles['Title']
    )
    elements.append(title)
    elements.append(Spacer(1, 0.2 * inch))

    # Subtitle
    subtitle = Paragraph(
        f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Total Products: {len(products)}",
        styles['Normal']
    )
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3 * inch))

    # Table data
    table_data = [["SKU", "Product", "Category", "Cost", "Price", "Stock", "Reorder", "Value", "Status"]]

    for product in products:
        stock_value = round(product.current_stock * product.unit_price, 2)
        if product.current_stock == 0:
            status = "Out of Stock"
        elif product.current_stock <= product.reorder_level:
            status = "Low Stock"
        else:
            status = "In Stock"

        table_data.append([
            product.sku,
            product.name[:25],
            product.category.name[:12] if product.category else "N/A",
            f"Rs.{product.unit_price}",
            f"Rs.{product.selling_price}",
            str(product.current_stock),
            str(product.reorder_level),
            f"Rs.{stock_value}",
            status
        ])

    # Total row
    total_value = round(sum(p.current_stock * p.unit_price for p in products), 2)
    table_data.append(["", "TOTAL", "", "", "", "", "", f"Rs.{total_value}", ""])

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        # Header
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4F46E5')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#F5F3FF')]),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#E5E7EB')),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F3F4F6')),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
    ]))

    elements.append(table)
    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def export_stock_movements_excel(db: Session) -> bytes:
    movements = db.query(StockMovement).order_by(
        StockMovement.created_at.desc()
    ).limit(500).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Movements"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="059669", end_color="059669", fill_type="solid")

    headers = ["Date & Time", "Product", "SKU", "Movement Type", "Quantity", "Recorded By", "Note"]
    col_widths = [20, 25, 15, 15, 10, 20, 30]

    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, movement in enumerate(movements, 2):
        row_data = [
            movement.created_at.strftime("%Y-%m-%d %H:%M"),
            movement.product.name if movement.product else f"Product #{movement.product_id}",
            movement.product.sku if movement.product else "N/A",
            movement.movement_type.value,
            movement.quantity,
            movement.user.name if movement.user else f"User #{movement.user_id}",
            movement.note or ""
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = Alignment(horizontal="center")

            if col == 4:
                if movement.movement_type.value == "STOCK_IN":
                    cell.font = Font(color="059669", bold=True)
                elif movement.movement_type.value == "STOCK_OUT":
                    cell.font = Font(color="D97706", bold=True)
                else:
                    cell.font = Font(color="2563EB", bold=True)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()